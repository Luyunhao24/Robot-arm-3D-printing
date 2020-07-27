import urx
import numpy as np
import serial , time
import socket
from estimate_plane import estimate_plane
from Coordinate_transformation import Coordinate_transformation
from openpyxl import load_workbook

wb = load_workbook(filename=r'C:\Users\LU\Desktop\test_1.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
arr1 = []
arr2 = []
arr3 = []
arr = []
for row_A in range(1, 210000):
    a1 = ws.cell(row=row_A, column=1).value
    a2 = ws.cell(row=row_A, column=2).value
    a3 = ws.cell(row=row_A, column=3).value
    if a1:
        arr1.append(a1)
    if a2:
        arr2.append(a2)
    if a3:
        arr3.append(a3)

arr = np.vstack((arr1,arr2,arr3))   #读取路径
N = arr.shape[1] - 1
t = []
a = []
b = []
t.append(10)
for i in range(N):  #计算路径时间
    s = ((arr[0][i+1]-arr[0][i])**2 + (arr[1][i+1]-arr[1][i])**2 + (arr[2][i+1]-arr[2][i])**2)**0.5
    clock = (s/0.01)+0.1
    t.append(clock)
    
R= [[ 0.67404229,-0.73868061,0.00423645,-0.21223078],[0.7386522,0.6740543,0.00661261,-0.30540021],[-0.0077402,-0.00132791,0.99996916,0.03895087], [0,0,0,1]] #R为转换矩阵

rob = urx.Robot("192.168.1.102")
rob.set_tcp((0.00006,-0.00065,0.12323,0,0,0))
rob.set_payload(0.5, (0,0,0))
print(arr.shape[1])

sum = 0#打印时间
for i in range(len(t)): 
    sum = sum + t[i]
Progress = 0
for i in range(0,arr.shape[1]):
    print(i)
    f = np.append(arr[:,i],[1],0).reshape(4,1)
    #print(f)
    r = np.matmul(R,f)
    r = np.delete(r,3,0)
    r = np.append(r,[[-3],[-0.811891],[-0.031571]],0)
    r = r.reshape(1,6)
    #print(r)
    move=r.tolist()
    #print(move[0])
    move = ','.join(map(str,move[0]))
    programString = "movel(p["+move+"],a=1.2,v=0.01)"
    #s.send(bytes(programString, encoding="utf8"))
    print(programString)
    rob.send_program(programString)
    Progress = Progress + t[i]
    ProgressBar = "(%f)/(%f)" %(Progress,sum)
    print(ProgressBar)
    time.sleep(t[i])