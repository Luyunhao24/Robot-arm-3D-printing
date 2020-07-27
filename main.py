import urx
import numpy as np
import serial , time
import socket
from estimate_plane import estimate_plane
from Coordinate_transformation import Coordinate_transformation
from openpyxl import load_workbook

wb = load_workbook(filename=r'C:\Users\LU\Desktop\test_circle.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
arr1 = []
arr2 = []
arr3 = []
arr = []
for row_A in range(1, 210000):
    a1 = ws.cell(row=row_A, column=1).value
    a2 = ws.cell(row=row_A, column=2).value
    a3 = ws.cell(row=row_A, column=3).value
    if a1:
        arr1.append(a1)
    if a2:
        arr2.append(a2)
    if a3:
        arr3.append(a3)

arr =np.vstack((arr1,arr2,arr3))   #读取路径

N = 3
points = np.ones((N,6))   #采集点N》=3

t = []
t.append(10)
for i in range(N):  #计算路径时间
    s = ((arr[0][i+1]-arr[0][i])**2 + (arr[1][i+1]-arr[1][i])**2 + (arr[2][i+1]-arr[2][i])**2)**0.5
    clock = (s/0.01)+0.1 #v=10mm/s
    t.append(clock)

rob = urx.Robot("192.168.1.102")   #连接机械臂
rob.set_tcp((-0.0022,0.00078,0.12350,0,0,0))
rob.set_payload(0.5, (0,0,0))

for i in range(N):  #采集点
    input("Waiting to get coordinates:")
    pose = rob.getl()  
    points[i,:] = pose

alpha = points[0,3]
beta = points[0,4]
gamma = points[0,5]
points = np.delete(points,[3,4,5],axis=1)
print(points)

[a,b,c,d]=estimate_plane(points)
R=Coordinate_transformation(points,a,b,c,d)  #计算转换矩阵

#HOST = "169.254.204.33" # TCP/IP连接机械臂
#PORT = 30002 
#s = socket.socket(socket.AF_INET, socket.SOCK_STREAM) 
#s.connect((HOST, PORT)) 
#s.send(b'set_payload(0.5)')
#s.send(b'set_tcp((-0.03753,0.03201,0.10609,0,0,0))')

#ser = serial.Serial()  #连接Arduino
#ser.baudrate = 250000
#ser.port = 'COM4'
#ser.timeout = 1
#ser.open()
#print(ser.isOpen())
#print(ser)
#print(ser.readlines())
#print(ser.in_waiting) 

#w = ser.write('M109 S200\n'.encode())  #等待喷头升温
#time.sleep(1)  #time
#while ser.in_waiting!=0:
    #print(ser.readline())
    #time.sleep(1)
    
sum = 0  #打印时间
for i in range(len(t)): 
    sum = sum + t[i]
Progress = 0
    
for i in range(0,arr.shape[1]):    #发送控制代码
    f = np.append(arr[:,i],[1],0).reshape(4,1)
    print(f)
    r = np.matmul(R,f)
    r = np.delete(r,3,0)
    r = np.append(r,[[alpha],[beta],[gamma]],0)
    r = r.reshape(1,6)
    print(r)
    move=r.tolist()
    print(move[0])
    move = ','.join(map(str,move[0]))
    programString = "movej(p["+move+"],a=1.4,v=1.04)" 
    print(programString)
    #s.send(bytes(programString, encoding="utf8")) #TCP/IP
    #rob.movel(move[0], acc=1.4, vel=1.04) #urx
    rob.send_program(programString) 
    Progress = Progress + t[i]
    ProgressBar = "(%f)/(%f)" %(Progress,sum)
    print(ProgressBar)  #打印进度
    time.sleep(t[i])
    
